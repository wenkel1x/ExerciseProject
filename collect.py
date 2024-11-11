
from datetime import datetime
from openpyxl import load_workbook
import os
import pandas as pd
import glob
import re
import json
import sys
#set factory different model
class base():
    def __init__(self,read_file,export_file):
        self.read_file = read_file
        self.export_file = export_file
        self.model_name = self.__class__.__name__
        #self.exec_init(data)
        #month_list = self.get_month_abbr()
        
    def exec_init(self,data):
        currency_month_abbr = datetime.now().strftime("%b")
        self.rack_fields = data.get(self.model_name,{}).get("rack_key_word",currency_month_abbr) or currency_month_abbr
        self.server_fields = data.get(self.model_name,{}).get("server_key_word",f"{currency_month_abbr} Server QTY") or "{} Server QTY".format(currency_month_abbr)
        self.mb_fields = data.get(self.model_name,{}).get("MB_key_word",f"{currency_month_abbr} MB QTY") or "{} MB QTY".format(currency_month_abbr)
        self.sb_fields = data.get(self.model_name,{}).get("SB_Key_word",f"{currency_month_abbr} SB QTY") or "{} SB QTY".format(currency_month_abbr)
        self.export_file_index = data.get("export_loading_index")
        self.get_after_num = data.get("get_after_num",5) or 5
        self.output_sheet_name = data.get("output_sheet_name","Loading by model") or "Loading by model"
        self.model_fac = data.get("fac_name",["QMF","QMN","QCG"]) or ["QMF","QMN","QCG"]
        self.col = self.get_excel_col()
        self.read_excel()

    def read_excel(self):
        sheet_name = pd.ExcelFile(self.read_file).sheet_names
        for fac in self.model_fac:
            pattern = re.compile(fr'.*{fac}.*',re.IGNORECASE)
            match_sheet = [name for name in sheet_name if pattern.match(name)]
            if not match_sheet:
                print(f"not exit sheet {fac} in {self.read_file} excel")
                continue
            df = pd.read_excel(self.read_file, sheet_name=match_sheet[0])
            last_row = df.tail(1)
            #get columns name list
            columns_index  = df.columns.tolist()
            if fac == "QMF" and self.model_name == "FAB":
                fa_index = columns_index.index(self.rack_fields)
                fa_data = last_row.iloc[:,fa_index:fa_index+self.get_after_num]  
                self.export_to_excel(fa_data,self.export_file_index[fac][self.model_name]["fa"])
                last_row = df.tail(2).head(1)
            rack_index = columns_index.index(self.rack_fields)
            server_index = columns_index.index(self.server_fields)
            mb_index = columns_index.index(self.mb_fields)
            sb_index = columns_index.index(self.sb_fields)
            
            rack_data = last_row.iloc[:,rack_index:rack_index+self.get_after_num]
            server_data = last_row.iloc[:,server_index:server_index+self.get_after_num]
            mb_data = last_row.iloc[:,mb_index:mb_index+self.get_after_num]
            sb_data = last_row.iloc[:,sb_index:sb_index+self.get_after_num]
            
            self.export_to_excel(rack_data,self.export_file_index[fac][self.model_name]["rack"])
            self.export_to_excel(server_data,self.export_file_index[fac][self.model_name]["server"])
            self.export_to_excel(mb_data,self.export_file_index[fac][self.model_name]["mb"])
            self.export_to_excel(sb_data,self.export_file_index[fac][self.model_name]["sb"])

    def get_month_abbr(self):
        now = datetime.now()
        month_abbr = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        current_month_index=now.month - 2
        export_month_index=[(current_month_index + i ) % 12 for i in range(self.get_after_num)]
        month_list = [month_abbr[i] for i in export_month_index]
        return month_list
    
    def get_excel_col(self):
        workbook = load_workbook(self.export_file)
        worksheet = workbook[self.output_sheet_name]
        for cell in worksheet.iter_rows(min_row=1,max_row=1,values_only=True):
            for index, value in enumerate(cell):
                if value == self.rack_fields:
                    col = index
            if col is not None:
                break
        return col

    def export_to_excel(self,data,row):
        df = pd.DataFrame(data)
        with pd.ExcelWriter(self.export_file, mode='a', engine='openpyxl',if_sheet_exists='overlay') as writer:
            df.to_excel(writer,sheet_name=self.output_sheet_name,index=False,header=False,startrow=row-1,startcol=self.col)
    

class FAB(base):
    def __init__(self,read_file,export_file):
        super().__init__(read_file,export_file)
class AMA(base):
    def __init__(self,read_file,export_file):
        super().__init__(read_file,export_file)
class MSF(base):
    def __init__(self,read_file,export_file):
        super().__init__(read_file,export_file)
class QCT(base):
    def __init__(self,read_file,export_file):
        super().__init__(read_file,export_file)
class NTA(base):
    def __init__(self,read_file,export_file):
        super().__init__(read_file,export_file)
class SEL(base):
    def __init__(self,read_file,export_file):
        super().__init__(read_file,export_file)

def main():
    get_cwd = os.getcwd()
    all_files = glob.glob('*.xlsx')
    pattern = re.compile(r'^(fab|ama|msf|qct|nta|sel).*\.xlsx$',re.IGNORECASE)
    loading_pattern = re.compile(r'.*loading.*\.xlsx$',re.IGNORECASE)
    match_file = [file for file in all_files if pattern.match(file) and not file.startswith('~')]
    if not match_file:
        raise ValueError("No excel files matching any words start with [fab ama msf qct nta sel] were found in {} path".format(get_cwd))
    loading_file_name = [file for file in all_files if loading_pattern.match(file) and not file.startswith('~')]
    if not loading_file_name[0]:
        raise ValueError("No loading field is included, Can't find any export data file in {} path".format(get_cwd))
    loading_file_path = os.path.join(get_cwd,loading_file_name[0])
    jsondata={}
    json_path=os.path.join(get_cwd,"collect.json")
    try:
        with open(json_path,'r') as file:
            jsondata = json.load(file)
    except json.JSONDecodeError:
        print("JSON file Decoding Error")
    except FileNotFoundError:
        print("can't find JSON file")
    except Exception as e:
        print(f"error: {e}")
    model_name=jsondata.get("model_name",["FAB","MSF","AMA","QCT","NTA","SEL"]) or ["FAB","MSF","AMA","QCT","NTA","SEL"]
    for file in match_file:
        for model in model_name:
            if model in file:
                print("collect {} data from file {}".format(model,file))
                try:
                    cls = getattr(sys.modules[__name__],model)
                    fac_work = cls(os.path.join(get_cwd,file),loading_file_path)
                    fac_work.exec_init(jsondata)
                except AttributeError:
                    print(f"Class {model} not found")
                break
    print("done")
if __name__ == '__main__':
    main()
